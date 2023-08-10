#Modulos
from tkinter import messagebox
import psycopg2
import openpyxl

#Funciones
def est_conexion():
    try:
        conn = psycopg2.connect(      
           bname = "siset", #base de datos
           user = "postgres", #usuario
           password = "Asea2023", #contraseña
           host = "localhost", #host
           port = "5432" #pueto de conexion 
        )
        messagebox.showinfo("conexion exitosa","conexion establecida")#Dialogo para saber si se complemento la conexion
    except Exception as ex:
        messagebox.showinfo("error en la conexion",ex)#dialogo de error
    finally:
        conn.close()#Cierre de la coneccion

#funcion de insertar de datos
def Insertar_datos(ruta):
    try:
        conn = psycopg2.connect(      
           dbname = "siset", #base de datos
           user = "postgres", #usuario
           password = "Asea2023", #contraseña
           host = "localhost", #host
           port = "5432" #pueto de conexion 
        )
        workbook = openpyxl.load_workbook(ruta) # lectura de archivo
        sheet = workbook.active#lee la hoja activa del documento
        filas = sheet.max_row + 1
        
        for x in range(2, filas):
            
            # BITÁCORA
            bitacora = str.format(sheet['B' + str(x) + ''].value)
            # RFC
            rfc = str.format(sheet['D' + str(x) + ''].value)
            

            cursor = conn.cursor()
            cursor.execute("INSERT INTO seguimiento (fsolicitud,bitacora_expediente, rnomrazonsolcial, rfc) VALUES (" + "'"+bitacora+"'" + ", "+"'"+rfc+"'"+");")
            conn.commit()
        messagebox.showinfo("Datos insertados","Sean insertado los datos con exito")#Mensaje de exito
    except Exception as ex:
        messagebox.showwarning("A ocurrido algo inesperado",ex)#dialogo de error 
    finally:
        conn.close()#Cierre de la conexion

def Actualizar_datos(ruta):
    try:
        conn = psycopg2.connect(      
           dbname = "siset",
           user = "postgres",
           password = "Asea2023",
           host = "localhost",
           port = "5432"
        )
        workbook = openpyxl.load_workbook(ruta) # lectura de archivo
        sheet = workbook.active #Lectura de hoja
        filas = sheet.max_row + 1 
        for x in range(2, filas):
            # BITÁCORA
            bitacora = str.format(sheet['B' + str(x) + ''].value)
            # RFC
            rfc = str.format(sheet['D' + str(x) + ''].value)
            
            cursor = conn.cursor()
            cursor.execute("UPDATE seguimiento SET rfc = '"+ rfc +"' Where bitacora_expediente = '"+bitacora+"'"+";")
            conn.commit()
        messagebox.showinfo("Datos Actualizados","Sean Actualizado los datos con exito")#Mensaje de exito
    except Exception as ex:
        messagebox.showwarning("A ocurrido algo inesperado",ex)#dialogo de error 
    finally:
        conn.close()#Cierre de la conexion